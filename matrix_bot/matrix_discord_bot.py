import discord
from discord_token_matrix import token #여기서 discord token 불러옵니다
from discord.ext import commands
import random
from 메뉴추천 import 메뉴리스트중복없음
import time
import pandas as pd
import openpyxl as op
import asyncio

bot = commands.Bot(command_prefix='!',intents=discord.Intents.all())
global TorFalse
TorFalse = 0
운영진 = ["최준수","음현식","이유찬"]
@bot.event
async def on_ready():
    game = discord.Game("명령어 <-- 입력으로 명령어보기")
    await bot.change_presence(status=discord.Status.online, activity=game)
    print("준비완료!")
    
@bot.event
async def on_message(message):
    #야옹answering
    if message.author.bot:
        return None
    else:
        if message.content == "야옹":
            await message.channel.send("야옹")
        if message.content[:2] == "야옹" and message.content[3:].isnumeric() == True:
            if int(message.content[3:]) <= 1000:
                if message.author.bot:
                    return None
                else:
                    await message.channel.send("야옹"*int(message.content[3:]))
            elif int(message.content[3:]) > 1000:
                await message.channel.send("1000 이하로 입력해줘 너무 길어..옹")

    #매냥아사랑해answering
    if message.content == "매냥아사랑해":
        await message.channel.send("나도 사랑해 야옹~ :heart:")
    
    #메뉴추천
    if message.content == "저메추":
        await message.channel.send(f"매냥이는 {random.choice(메뉴리스트중복없음)} (을)를 추천해! 맛있게 먹냥 :heart:")

    #주사위 굴리기
    if message.content == "주사위":
        await message.channel.send("https://tenor.com/view/dice-gif-18958117")
        await message.channel.send("주사위 굴러가는중...")
        time.sleep(2)
        random_num = random.randint(1,6)
        await message.channel.send(f"{message.author.display_name}의 주사위 결과는 : {random_num} 이야! :heart:")

    #help
    if message.content == "명령어":
        #커맨드 손수 추가
        command = "저메추, 주사위, 전부집합, 돌림판, 그럼제가선배맘에, -나를 속인거니?, 매냥아"
        await message.channel.send(f"현재 등록된 명령어는 [ {command} ] 가 있어! 언제든지 쓰고싶으면 얘기해죠! :heart: \n참고로 [전부집합] 명령어는 everyone 멘션이기 때문에 사용에 각별히 주의해줘!!")

    #@everyone 치기
    if message.content == "전부집합":
        await message.channel.send(f"@everyone {message.author.display_name}(이)가 모두를 집합시켰어!!")

    #돌림판
    if message.content[:3] == "돌림판":
        돌림판리스트 = []
        돌림판리스트 = message.content[4:].split()
        await message.channel.send(f"돌림판 결과는 {돌림판리스트[random.randint(0,len(돌림판리스트))]} (이)에요!!")

    #그럼제가선배맘에
    if message.content == "그럼제가선배맘에":
        await message.channel.send("탕탕후루후루 타탕탕 후루루루루!!!!!")
        await message.channel.send("https://tenor.com/view/triples-kotone-s11-malatanghulu-malatang-gif-981579389405623666")

    
    #짱구야나를속인거니?
    if message.content[-8:] == "나를 속인거니?":
        await message.channel.send(file = discord.File("fakeme.webp"))

    #대답
    if message.content == "매냥아":
        대답 = ["네?", "매냥이 여기있어요!", "듣고있어요!", "무슨 일이신가요?", "야옹~"]
        대답_선택num = random.randint(0,4)
        await message.channel.send(대답[대답_선택num])

    #출석체크
    file_name = "matrix_bot\m_registrations.xlsx"
    wb = op.load_workbook(file_name)
    ws = wb.active
    

    def 운영진확인(x):
        if x in 운영진:
            return True
        else:
            return False

    def check(m):
        return m.author == message.author and m.channel == message.channel

    if message.content == "출석체크시작":
        if 운영진확인(message.author.display_name) == True:
            global 출석체크y
            global 출석체크x
            출석체크y = 1
            await message.channel.send("몇주차인가요? 숫자로 입력해주세요!")
            try:
                출석체크x = await bot.wait_for("message",check = check,timeout=5)
                출석체크x = 출석체크x.content
                global TorFalse
                TorFalse = 1
                await message.channel.send(f"@everyone 매트릭스 {출석체크x}주차 출석체크가 시작되었습니다!")
            except asyncio.TimeoutError:
                await message.channel.send("시간초과에요!")
                return None

    def 출석체크이름찾기(x):
        global 출석체크y
        if x != ws.cell(row=출석체크y, column=1).value:
            출석체크y = 출석체크y + 1
            출석체크이름찾기(x)
        elif ws.cell(row=출석체크y, column=1).value == temp출석체크이름:
            global 출석체크x
            ws.cell(row=출석체크y, column=(int(출석체크x)+1)).value = "O"
            wb.save(file_name)

    if message.content == "출석체크":
        if TorFalse == 1:
            global temp출석체크이름
            temp출석체크이름 = message.author.display_name
            출석체크이름찾기(temp출석체크이름)
            await message.channel.send(f"{message.author.display_name}님 출석체크가 완료되었습니다!")
        else:
            await message.channel.send("출석체크 가능 시간이 아닙니다")

    if message.content == "출석체크종료" and 운영진확인(message.author.display_name):
        TorFalse = 0
        await message.channel.send("@everyone 매트릭스 출석체크가 종료되었습니다!")

bot.run(token)